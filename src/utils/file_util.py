#!/usr/local/bin/python3
# -*- coding: utf-8 -*-
"""

"""
import json
import os
import csv
import openpyxl
import requests
import yaml
import xml.etree.ElementTree as ET


def read_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    file_extension = file_extension.lower()

    with open(file_path, 'r', encoding='utf-8') as f:
        if file_extension == '.csv':
            return list(csv.reader(f))
        elif file_extension == '.tsv':
            return list(csv.reader(f, delimiter='\t'))
        elif file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            return [[cell.value for cell in row] for row in sheet.iter_rows(values_only=True)]
        elif file_extension in ['.yaml', '.yml']:
            return yaml.safe_load(f)
        elif file_extension == '.xml':
            return ET.parse(f)
        else:
            return f.read()


def write_to_file(file_path, file_data):
    """将字典或列表数据处理为Excel、文本、CSV、YAML、XML、YML和JSON文件

    Args:
        file_data (dict or list): 要处理的数据
        file_path (str): 要保存的文件路径，包括文件名和扩展名

    Returns:
        bool: 操作是否成功
    """
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension == '':
        print("Invalid file path: {}".format(file_path))
        return False
    if file_extension not in ['.xlsx', '.txt', '.csv', '.yaml', '.yml', '.xml', '.json']:
        print("Unsupported file type: {}".format(file_extension))
        return False

    file_type = file_extension[1:]

    try:
        if file_type == 'excel':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in file_data:
                sheet.append(row)
            workbook.save(file_path)
        elif file_type == 'txt':
            with open(file_path, 'w') as file:
                for row in file_data:
                    file.write(str(row) + '\n')
        elif file_type == 'csv':
            with open(file_path, 'w', newline='') as file:
                writer = csv.writer(file)
                for row in file_data:
                    writer.writerow(row)
        elif file_type in ['yaml', 'yml']:
            with open(file_path, 'w') as file:
                yaml.dump(file_data, file)
        elif file_type == 'xml':
            root = ET.Element('root')
            for row in file_data:
                item = ET.SubElement(root, 'item')
                for key, value in row.items():
                    ET.SubElement(item, key).text = str(value)
            tree = ET.ElementTree(root)
            tree.write(file_path)
        elif file_type == 'json':
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump(file_data, file)
    except Exception as e:
        print("Failed to write data to file: {}".format(str(e)))
        return False

    return True


def read_ol_file(_url):
    response = requests.get(_url)
    print("====response:", response)
    html_content = response.text
    print("====html_content:", html_content)
    return html_content


if __name__ == '__main__':
    path = r"D:\无标题表格(2).tsv"
    data = read_file(path)
    print(data)
